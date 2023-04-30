import os
from datetime import datetime
from PyQt5 import QtCore, QtGui, QtWidgets, uic
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QMessageBox, QMainWindow, QStyledItemDelegate, QAction
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import pandas as pd
import common as cm
from database import connect_database
from source import add_admin
from source import edit_admin
from source import login_3
from PyQt5.uic import loadUi

data = connect_database.select_data_user()


class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        return


class UI_admin(QMainWindow):
    def __init__(self, parent=None):
        super(UI_admin, self).__init__(parent)
        # self.database_config_obj = None
        # uic.loadUi('admin.ui', self)

    def setupUi(self):
        self.setObjectName("MainWindow")
        self.resize(1010, 618)
        self.centralwidget = QtWidgets.QWidget(self)
        self.setWindowIcon(QtGui.QIcon("../icon/logo.png"))
        self.centralwidget.setObjectName("centralwidget")
        self.table_admin = QtWidgets.QTableWidget(self.centralwidget)
        delegate = ReadOnlyDelegate(self)
        self.table_admin.setItemDelegateForColumn(0, delegate)
        self.table_admin.setItemDelegateForColumn(1, delegate)
        self.table_admin.setItemDelegateForColumn(2, delegate)
        self.table_admin.setItemDelegateForColumn(3, delegate)
        self.table_admin.setItemDelegateForColumn(4, delegate)
        self.table_admin.setItemDelegateForColumn(5, delegate)
        self.table_admin.setItemDelegateForColumn(6, delegate)
        self.table_admin.setItemDelegateForColumn(7, delegate)
        self.table_admin.setGeometry(QtCore.QRect(30, 110, 951, 411))
        self.table_admin.setObjectName("table_admin")
        self.table_admin.setColumnCount(10)
        self.table_admin.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        self.table_admin.setHorizontalHeaderItem(9, item)
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(410, 10, 191, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.txt_search = QtWidgets.QLineEdit(self.centralwidget)
        self.txt_search.setGeometry(QtCore.QRect(390, 60, 461, 31))
        self.txt_search.setObjectName("txt_search")
        self.btn_search = QtWidgets.QPushButton(self.centralwidget)
        self.btn_search.setStyleSheet("background-color:rgb(94, 186, 125)")
        self.btn_search.clicked.connect(self.search)
        self.btn_search.setIcon(QtGui.QIcon('../icon/search.png'))
        self.btn_search.setGeometry(QtCore.QRect(890, 60, 89, 31))
        self.btn_search.setObjectName("btn_search")
        self.btn_add = QtWidgets.QPushButton(self.centralwidget)
        self.btn_add.setStyleSheet("background-color:rgb(94, 186, 125)")
        self.btn_add.setGeometry(QtCore.QRect(890, 530, 89, 41))
        self.btn_add.setObjectName("btn_add")
        self.btn_add.clicked.connect(self.add)
        self.btn_add.setIcon(QtGui.QIcon('../icon/add.png'))
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(840, 0, 81, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(920, 0, 55, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(50, 90, 51, 16))
        self.label_4.setObjectName("label_4")
        self.number = QtWidgets.QLabel(self.centralwidget)
        self.number.setGeometry(QtCore.QRect(90, 90, 31, 16))
        self.number.setObjectName("number")
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(110, 90, 55, 16))
        self.label_6.setObjectName("label_6")
        self.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1010, 25))
        self.menubar.setObjectName("menubar")
        self.menuFile = QtWidgets.QMenu(self.menubar)
        self.menuFile.setObjectName("menuFile")
        self.menuOption = QtWidgets.QMenu(self.menubar)
        self.menuOption.setObjectName("menuOption")
        self.menuOption_2 = QtWidgets.QMenu(self.menubar)
        self.menuOption_2.setObjectName("menuOption_2")
        self.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName("statusbar")
        self.setStatusBar(self.statusbar)
        self.actionSign_out = QtWidgets.QAction(self)
        self.actionSign_out.setObjectName("actionSign_out")
        self.actionExit = QtWidgets.QAction(self)
        self.actionExit.setIcon(QtGui.QIcon('../icon/sign-out.png'))
        self.actionExit.setObjectName("actionExit")
        self.actionExport_to_excel = QtWidgets.QAction(self)
        self.actionExport_to_excel.triggered.connect(self.export_button_active)
        self.actionExport_to_excel.setIcon(QtGui.QIcon('../icon/export.png'))
        self.actionExport_to_excel.setObjectName("actionExport_to_excel")
        self.actionAbout_us = QtWidgets.QAction(self)
        self.actionAbout_us.setIcon(QtGui.QIcon('../icon/contact.png'))
        self.actionAbout_us.triggered.connect(self.about)
        self.actionAbout_us.setObjectName("actionAbout_us")
        self.actionSign_out_2 = QtWidgets.QAction(self)
        self.actionSign_out_2.triggered.connect(self.sign_out)
        self.actionSign_out_2.setObjectName("actionSign_out_2")
        self.actionExit_2 = QtWidgets.QAction(self)
        self.actionExit_2.setIcon(QtGui.QIcon('../icon/sign-out.png'))

        self.actionExit_2.triggered.connect(self.exit)
        self.actionExit_2.setObjectName("actionExit_2")
        self.menuFile.addAction(self.actionExport_to_excel)
        self.menuOption.addAction(self.actionAbout_us)
        self.menuOption_2.addAction(self.actionSign_out_2)
        self.menuOption_2.addAction(self.actionExit_2)
        self.menubar.addAction(self.menuFile.menuAction())
        self.menubar.addAction(self.menuOption.menuAction())
        self.menubar.addAction(self.menuOption_2.menuAction())
        self.retranslateUi(self)
        self.init_data(data)

        # QtCore.QMetaObject.connectSlotsByName(self)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        item = self.table_admin.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "ID"))
        item = self.table_admin.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Name"))
        item = self.table_admin.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Date of Birth"))
        item = self.table_admin.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "Phone"))
        item = self.table_admin.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Email"))
        item = self.table_admin.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "Address"))
        item = self.table_admin.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Username"))
        item = self.table_admin.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "Password"))
        item = self.table_admin.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "Edit"))
        item = self.table_admin.horizontalHeaderItem(9)
        item.setText(_translate("MainWindow", "Delete"))
        self.label_3.setText(_translate("MainWindow", "ADMINISTRATORS"))
        self.btn_search.setText(_translate("MainWindow", "Search"))
        self.btn_add.setText(_translate("MainWindow", "Add"))
        self.label.setText(_translate("MainWindow", "Welcome: "))
        self.label_2.setText(_translate("MainWindow", "Test"))
        self.label_4.setText(_translate("MainWindow", "Total:"))
        self.number.setText(_translate("MainWindow", "23"))
        self.label_6.setText(_translate("MainWindow", "User"))
        self.menuFile.setTitle(_translate("MainWindow", "File"))
        self.menuOption.setTitle(_translate("MainWindow", "About"))
        self.menuOption_2.setTitle(_translate("MainWindow", "Option"))
        self.actionSign_out.setText(_translate("MainWindow", "Sign out"))
        self.actionExit.setText(_translate("MainWindow", "Exit"))
        self.actionExport_to_excel.setText(_translate("MainWindow", "Export to excel"))
        self.actionAbout_us.setText(_translate("MainWindow", "About us"))
        self.actionSign_out_2.setText(_translate("MainWindow", "Sign out"))
        self.actionExit_2.setText(_translate("MainWindow", "Exit"))

    def init_data(self, data):
        try:
            total = connect_database.count_number()
            self.number.setText(str(total[0][0]))
            for i in range(0, len(data)):
                self.table_admin.insertRow(i)
                self.table_admin.setItem(i, 0, QtWidgets.QTableWidgetItem(str(data[i][0])))
                self.table_admin.setItem(i, 1, QtWidgets.QTableWidgetItem(str(data[i][1])))
                self.table_admin.setItem(i, 2, QtWidgets.QTableWidgetItem(str(data[i][2])))
                self.table_admin.setItem(i, 3, QtWidgets.QTableWidgetItem(str(data[i][3])))
                self.table_admin.setItem(i, 4, QtWidgets.QTableWidgetItem(str(data[i][4])))
                self.table_admin.setItem(i, 5, QtWidgets.QTableWidgetItem(str(data[i][5])))
                self.table_admin.setItem(i, 6, QtWidgets.QTableWidgetItem(str(data[i][6])))
                self.table_admin.setItem(i, 7, QtWidgets.QTableWidgetItem(str(data[i][7])))
                self.table_admin.setItem(i, 8, QtWidgets.QTableWidgetItem(str(data[i][7])))
                self.table_admin.setItem(i, 9, QtWidgets.QTableWidgetItem(str(data[i][7])))

                self.detail_btn = QtWidgets.QPushButton(self)
                self.detail_btn.clicked.connect(self.edit_btn_click)
                self.detail_btn.setText("Edit")
                self.table_admin.setCellWidget(i, 8, self.detail_btn)

                self.detail_btn = QtWidgets.QPushButton(self)
                self.detail_btn.clicked.connect(self.delete_btn_click)
                self.detail_btn.setText("Delete")
                self.table_admin.setCellWidget(i, 9, self.detail_btn)
        except:
            QMessageBox.information(self, 'System',
                                    'An error occurred while passing data to the table. Please try again!',
                                    QMessageBox.Close)

    def edit_btn_click(self):
        try:
            self.ui = QMainWindow()
            self.ui = edit_admin.UI_edit_admin()
            self.ui.setupUi()
            self.ui.show()
            button = QtWidgets.qApp.focusWidget()
            index = self.table_admin.indexAt(button.pos())
            table_model = self.table_admin.model()
            id_index = table_model.index(index.row(), 0)
            id = table_model.data(id_index)
            name_index = table_model.index(index.row(), 1)
            name = table_model.data(name_index)
            dob_index = table_model.index(index.row(), 2)
            dob = table_model.data(dob_index)
            phone_index = table_model.index(index.row(), 3)
            phone = table_model.data(phone_index)
            email_index = table_model.index(index.row(), 4)
            email = table_model.data(email_index)
            address_index = table_model.index(index.row(), 5)
            address = table_model.data(address_index)
            username_index = table_model.index(index.row(), 6)
            username = table_model.data(username_index)

            self.ui.txt_name_2.setText(id)
            self.ui.txt_name_3.setText(username)
            self.ui.txt_name_4.setText(name)
            date_object = datetime.strptime(dob, '%Y-%m-%d').date()
            self.ui.dateEdit.setDate(date_object)
            self.ui.txt_name_5.setText(email)
            self.ui.txt_name_6.setText(address)
            self.ui.btn_ok.clicked.connect(self.edit_admin)
            self.ui.btn_canncel.clicked.connect(self.cancel_edit_admin)
        except:
            QMessageBox.information(self, 'System',
                                    'An error occurred while editing information. Please try again later!',
                                    QMessageBox.Close)

    def cancel_edit_admin(self):
        try:
            self.ui = QMainWindow()
            self.ui = edit_admin.UI_edit_admin()
            self.ui.setupUi()
            self.ui.hide()
        except:
            QMessageBox.information(self, 'System', 'An error occurred while closing. Please try again!',
                                    QMessageBox.Close)

    def edit_admin(self):
        try:
            id = self.ui.txt_name_2.text()
            username = self.ui.txt_name_3.text()
            name = self.ui.txt_name_4.text()
            dob = self.ui.dateEdit.dateTime().toPyDateTime()
            email = self.ui.txt_name_5.text()
            address = self.ui.txt_name_6.text()
            if connect_database.edit_admin(id, name, dob, email, address, username):
                QMessageBox.information(self.ui, 'Message', f'Successfully edit the user: {username} ',
                                        QMessageBox.Close)
                self.table_admin.clearContents()
                self.table_admin.setRowCount(0)
                data_new = connect_database.select_data_user()
                self.init_data(data_new)

            else:
                QMessageBox.information(self.ui, 'Message', f'Failed edit the user: {username} ', QMessageBox.Close)
                self.table_admin.clearContents()
                self.table_admin.setRowCount(0)
                data_new = connect_database.select_data_user()
                self.init_data(data_new)
            self.ui = QMainWindow()
            self.ui = edit_admin.UI_edit_admin()
            self.ui.setupUi()
            self.ui.hide()
        except:
            QMessageBox.information(self, 'System',
                                    'An error occurred while editing information. Please try again later!',
                                    QMessageBox.Close)

    def delete_btn_click(self):
        try:
            button = QtWidgets.qApp.focusWidget()
            index = self.table_admin.indexAt(button.pos())
            table_model = self.table_admin.model()
            id_index = table_model.index(index.row(), 0)
            id = table_model.data(id_index)
            name_index = table_model.index(index.row(), 1)
            name = table_model.data(name_index)
            ques = QMessageBox.question(self, 'System', f'Are you sure you want to delete user: {name}?',
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ques == QMessageBox.Yes:
                if connect_database.delete_data(id):
                    QMessageBox.about(self, "System", f"Successfully delete the user: {name}")
                    self.table_admin.clearContents()
                    self.table_admin.setRowCount(0)
                    data_new = connect_database.select_data_user()
                    self.init_data(data_new)
                else:
                    QMessageBox.about(self, "System", "Delete failed, try again")
        except:
            QMessageBox.information(self, 'System',
                                    'An error occurred while editing information. Please try again later!',
                                    QMessageBox.Close)

    def search(self):
        try:
            search_text = self.txt_search.text()
            data = connect_database.search_user(search_text)
            if len(data) != 0:
                self.table_admin.clearContents()
                self.table_admin.setRowCount(0)
                self.init_data(data)
            else:
                QMessageBox.information(self, 'Search', 'Nothing found!', QMessageBox.Close)
        except:
            QMessageBox.information(self, 'System',
                                    'An error occurred while search user. Please try again later!',
                                    QMessageBox.Close)

    def add(self):
        try:
            self.ui = QMainWindow()
            self.ui = add_admin.UI_add_user()
            self.ui.setupUi()
            self.ui.show()
            self.ui.btn_ok.clicked.connect(self.ok)
            self.ui.btn_canncel.clicked.connect(self.cancel)
        except:
            QMessageBox.information(self, 'System',
                                    'An error occurred while add new user. Please try again later!',
                                    QMessageBox.Close)

    def ok(self):
        try:
            name = self.ui.txt_name.text()
            dob = self.ui.date.dateTime().toPyDateTime().date()
            phone = self.ui.txt_phone.text()
            email = self.ui.txt_email.text()
            address = self.ui.txt_email_2.text()
            username = self.ui.tx_username.text()
            password = self.ui.txt_password.text()
            role = self.ui.comboBox.currentText()
            date_now = datetime.now().date()
            if role == 'admin':
                role = str('1')
            else:
                role = str('0')
            if all(v != '' for v in [name, dob, phone, email, address, username, password, role]):
                if dob < date_now:
                    if cm.check_phone(phone):
                        if cm.check_email(email):
                            try:
                                if self.ui.check_box_agree.isChecked():
                                    if connect_database.register(name, dob, phone, email, address, username, password,
                                                                 role):
                                        QMessageBox.information(self.ui, 'Add user',
                                                                'Congratulations. Successful registration',
                                                                QMessageBox.Close)
                                        self.ui.hide()
                                        data_new = connect_database.select_data_user()
                                        self.table_admin.clearContents()
                                        self.table_admin.setRowCount(0)
                                        self.init_data(data_new)

                                else:
                                    QMessageBox.information(self.ui, 'Add user', 'Please click the agree button!',
                                                            QMessageBox.Close)
                            except:
                                QMessageBox.information(self.ui, 'Add user', 'Registration failed!', QMessageBox.Close)
                        else:
                            QMessageBox.information(self.ui, 'Add user', 'Wrong email address format!',
                                                    QMessageBox.Close)
                    else:
                        QMessageBox.information(self.ui, 'Add user', 'Wrong phone number format!', QMessageBox.Close)
                else:
                    QMessageBox.information(self.ui, 'Add user', 'Date of birth cannot be greater than current date!',
                                            QMessageBox.Close)
            else:
                QMessageBox.information(self.ui, 'Add user', 'Please fill in all fields!', QMessageBox.Close)
        except:
            QMessageBox.information(self.ui, 'System',
                                    'An error occurred while registering. Please try again later!',
                                    QMessageBox.Close)

    def cancel(self):
        self.hide()

    def export(self):
        try:
            columnHeaders = []
            for j in range(self.table_admin.model().columnCount() - 2):
                columnHeaders.append(self.table_admin.horizontalHeaderItem(j).text())
            df = pd.DataFrame(columns=columnHeaders)
            for row in range(self.table_admin.rowCount()):
                for col in range(self.table_admin.columnCount() - 2):
                    df.at[row, columnHeaders[col]] = self.table_admin.item(row, col).text()
            wb = Workbook()
            wb = load_workbook('../Template/Template_user_export.xlsx')
            ws1 = wb.worksheets[0]
            offset_row = 1
            offset_col = 0
            row = 1
            for row_data in dataframe_to_rows(df, index=False, header=False):
                col = 1
                for cell_data in row_data:
                    ws1.cell(row + offset_row, col + offset_col, cell_data)
                    col += 1
                row += 1
            if os.path.exists('../Output') == False:
                folder_path = '../Output'
                os.mkdir(folder_path)
            (d, m, y, h, mi, s) = cm.split_date_time()
            wb.save(
                f'../Output/user_{d}_{m}_{y}_{h}_{mi}_{s}.xlsx')
        except:
            QMessageBox.information(self, 'System',
                                    'An error occurred while exporting to excel. Please try again later!',
                                    QMessageBox.Close)

    def data_from_table(self):
        try:
            row = self.table_admin.rowCount()
            df = pd.DataFrame(
                columns=['ID', 'Name', 'Date of Birth', 'Phone', 'Email', 'Address', 'Password'])
            for i in range(row):
                value = []
                for j in range(7):
                    value.append(self.table_admin.item(i, j).text())
                df.loc[len(df.index)] = value
            return df
        except Exception as err:
            QMessageBox.information(self, 'Error', str(err.args[0]), QMessageBox.Close)
            return []

    def export_button_active(self):
        try:
            df = self.data_from_table()

            if len(df.index) == 0:
                QMessageBox.information(self, 'Error', 'There is no value', QMessageBox.Close)
                return
            if cm.data_to_excel('User', df):
                QMessageBox.information(self, 'Success', 'Export success full', QMessageBox.Close)
            else:
                QMessageBox.information(self, 'Fail', 'Something wrong', QMessageBox.Close)
        except Exception as err:
            QMessageBox.information(self, 'Error', str(err.args[0]), QMessageBox.Close)

    def exit(self):
        try:
            ques = QMessageBox.question(self, 'System', 'Are you sure you want to exit the administrator?',
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ques == QMessageBox.Yes:
                self.hide()
                self.ui = QMainWindow()
                self.ui = login_3.UI_login()
                self.ui.setupUi()
                self.ui.show()
            else:
                pass
        except:
            QMessageBox.information(self, 'System', 'An error occurred while closing. Please try again!',
                                    QMessageBox.Close)

    def about(self):
        pass

    def sign_out(self):
        pass

    def closeEvent(self, event):
        try:
            ques = QMessageBox.question(self, 'Close System', 'Are you sure you want to exit system?',
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ques == QMessageBox.Yes:
                self.hide()
                self.ui = QMainWindow()
                self.ui = login_3.UI_login()
                self.ui.setupUi()
                self.ui.show()
            else:
                event.ignore()
        except:
            QMessageBox.information(self, 'System', 'An error occurred while closing. Please try again!',
                                    QMessageBox.Close)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ui = UI_admin()
    ui.setupUi()
    ui.show()
    sys.exit(app.exec_())
