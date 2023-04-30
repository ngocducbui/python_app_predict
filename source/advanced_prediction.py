import os
import pathlib
import re

import pandas as pd
import xlsxwriter
from PyQt5.QtWidgets import QMessageBox, QMainWindow, QStyledItemDelegate, QAction, QLineEdit
import sys

from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from xlsxwriter import Workbook

import machine_learning
from source import predict
from PyQt5 import QtCore, QtGui, QtWidgets
import tkinter as tk
from tkinter import filedialog
import common as cm
from sklearn.metrics import accuracy_score


class UI_advanced_prediction(QMainWindow):
    def __init__(self, max_depth, number, parent=None):
        self.max_depth = max_depth
        self.number = number
        self.list_data = []
        self.list_data_predict = []
        self.link_export = ''
        self.accuracy = 0
        super(UI_advanced_prediction, self).__init__(parent)

    def setupUi(self):
        self.setObjectName("MainWindow")
        self.resize(1122, 579)
        self.centralwidget = QtWidgets.QWidget(self)
        self.setWindowIcon(QtGui.QIcon("../icon/logo.png"))
        self.centralwidget.setObjectName("centralwidget")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(430, 0, 321, 51))
        font = QtGui.QFont()
        font.setFamily("Segoe UI Light")
        font.setPointSize(22)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_3.setFont(font)
        self.label_3.setStyleSheet("color: rgb(11, 34, 57);")
        self.label_3.setObjectName("label_3")
        self.data_table = QtWidgets.QTableWidget(self.centralwidget)
        self.data_table.setGeometry(QtCore.QRect(20, 60, 1081, 451))
        self.data_table.setObjectName("data_table")
        self.data_table.setColumnCount(35)
        self.data_table.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(0, 0, 0))
        self.data_table.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(10, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(11, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(12, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(13, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(14, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(15, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(16, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(17, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(18, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(19, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(20, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(21, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(22, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(23, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(24, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(25, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(26, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(27, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(28, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(29, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(30, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(31, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(32, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(33, item)
        item = QtWidgets.QTableWidgetItem()
        self.data_table.setHorizontalHeaderItem(34, item)

        self.import_btn = QtWidgets.QPushButton(self.centralwidget)
        self.import_btn.setStyleSheet("background-color:rgb(94, 186, 125)")
        self.import_btn.clicked.connect(self.import_csv)
        self.import_btn.setIcon(QtGui.QIcon('../icon/import.png'))
        self.import_btn.setGeometry(QtCore.QRect(990, 20, 111, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.import_btn.setFont(font)
        self.import_btn.setObjectName("import_btn")
        self.btn_close = QtWidgets.QPushButton(self.centralwidget)
        self.btn_close.setStyleSheet("background-color:rgb(218, 91, 11)")
        self.btn_close.clicked.connect(self.close)
        self.btn_close.setIcon(QtGui.QIcon('../icon/cancel.png'))
        self.btn_close.setGeometry(QtCore.QRect(990, 520, 111, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_close.setFont(font)
        self.btn_close.setObjectName("btn_close")
        self.btn_export = QtWidgets.QPushButton(self.centralwidget)
        self.btn_export.clicked.connect(self.export_button_active)
        self.btn_export.setIcon(QtGui.QIcon('../icon/export.png'))
        self.btn_export.setEnabled(False)
        self.btn_export.setStyleSheet("background-color:rgb(94, 186, 125)")
        self.btn_export.setGeometry(QtCore.QRect(860, 520, 121, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_export.setFont(font)
        self.btn_export.setObjectName("btn_export")
        self.label_link = QtWidgets.QLabel(self.centralwidget)
        self.label_link.hide()
        self.label_link.setGeometry(QtCore.QRect(120, 20, 55, 16))
        self.label_link.setObjectName("label_link")
        self.btn_compare = QtWidgets.QPushButton(self.centralwidget)
        self.btn_compare.setStyleSheet("background-color:rgb(94, 186, 125)")
        self.btn_compare.clicked.connect(self.compare)
        self.btn_compare.setIcon(QtGui.QIcon('../icon/compare.png'))
        self.btn_compare.setEnabled(False)
        self.btn_compare.setGeometry(QtCore.QRect(730, 520, 121, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_compare.setFont(font)
        self.btn_compare.setObjectName("btn_compare")
        self.btn_predict_all = QtWidgets.QPushButton(self.centralwidget)
        self.btn_predict_all.setStyleSheet("background-color:rgb(94, 186, 125)")
        self.btn_predict_all.clicked.connect(self.predict_all)
        self.btn_predict_all.setIcon(QtGui.QIcon('../icon/prediction.png'))
        self.btn_predict_all.setEnabled(False)
        self.btn_predict_all.setGeometry(QtCore.QRect(600, 520, 121, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btn_predict_all.setFont(font)
        self.btn_predict_all.setObjectName("btn_predict_all")
        self.label_tt = QtWidgets.QLabel(self.centralwidget)
        self.label_tt.hide()
        self.label_tt.setGeometry(QtCore.QRect(260, 20, 55, 16))
        self.label_tt.setObjectName("label_tt")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(40, 40, 101, 16))
        self.label.setObjectName("label")
        self.label_shape = QtWidgets.QLabel(self.centralwidget)
        self.label_shape.setGeometry(QtCore.QRect(140, 40, 81, 16))
        self.label_shape.setObjectName("label_shape")
        self.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1122, 25))
        self.menubar.setObjectName("menubar")
        self.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(self)
        self.statusbar.setObjectName("statusbar")
        self.setStatusBar(self.statusbar)
        self.retranslateUi()
        # QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("MainWindow", "Advanced Prediction"))
        self.label_3.setText(_translate("MainWindow", "Advanced Prediction"))
        item = self.data_table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Marital status"))
        item = self.data_table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Application mode"))
        item = self.data_table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Application order"))
        item = self.data_table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "Course"))
        item = self.data_table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "Daytime/evening attendance"))
        item = self.data_table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "Previous qualification"))
        item = self.data_table.horizontalHeaderItem(6)
        item.setText(_translate("main_window", "Nacionality"))
        item = self.data_table.horizontalHeaderItem(7)
        item.setText(_translate("main_window", "Mother\'s qualification"))
        item = self.data_table.horizontalHeaderItem(8)
        item.setText(_translate("main_window", "Father\'s qualification"))
        item = self.data_table.horizontalHeaderItem(9)
        item.setText(_translate("main_window", "Mother\'s occupation"))
        item = self.data_table.horizontalHeaderItem(10)
        item.setText(_translate("main_window", "Father\'s occupation"))
        item = self.data_table.horizontalHeaderItem(11)
        item.setText(_translate("main_window", "Displaced"))
        item = self.data_table.horizontalHeaderItem(12)
        item.setText(_translate("main_window", "Educational special needs"))
        item = self.data_table.horizontalHeaderItem(13)
        item.setText(_translate("main_window", "Debtor"))
        item = self.data_table.horizontalHeaderItem(14)
        item.setText(_translate("main_window", "Tuition fees up to date"))
        item = self.data_table.horizontalHeaderItem(15)
        item.setText(_translate("main_window", "Gender"))
        item = self.data_table.horizontalHeaderItem(16)
        item.setText(_translate("main_window", "Scholarship holder"))
        item = self.data_table.horizontalHeaderItem(17)
        item.setText(_translate("main_window", "Age at enrollment"))
        item = self.data_table.horizontalHeaderItem(18)
        item.setText(_translate("main_window", "International"))
        item = self.data_table.horizontalHeaderItem(19)
        item.setText(_translate("main_window", "Curricular units 1st sem (credited)"))
        item = self.data_table.horizontalHeaderItem(20)
        item.setText(_translate("main_window", "Curricular units 1st sem (enrolled)"))
        item = self.data_table.horizontalHeaderItem(21)
        item.setText(_translate("main_window", "Curricular units 1st sem (evaluations)"))
        item = self.data_table.horizontalHeaderItem(22)
        item.setText(_translate("main_window", "Curricular units 1st sem (approved)"))
        item = self.data_table.horizontalHeaderItem(23)
        item.setText(_translate("main_window", "Curricular units 1st sem (grade)"))
        item = self.data_table.horizontalHeaderItem(24)
        item.setText(_translate("main_window", "Curricular Units 1st Sem (without Evaluations)"))
        item = self.data_table.horizontalHeaderItem(25)
        item.setText(_translate("main_window", "Curricular Units 2nd Sem (credited)"))
        item = self.data_table.horizontalHeaderItem(26)
        item.setText(_translate("main_window", "Curricular Units 2nd Sem (enrolled)"))
        item = self.data_table.horizontalHeaderItem(27)
        item.setText(_translate("main_window", "Curricular Units 2nd Sem (evaluations)"))
        item = self.data_table.horizontalHeaderItem(28)
        item.setText(_translate("main_window", "Curricular Units 2nd Sem (approved)"))
        item = self.data_table.horizontalHeaderItem(29)
        item.setText(_translate("main_window", "Curricular Units 2nd Sem (grade)"))
        item = self.data_table.horizontalHeaderItem(30)
        item.setText(_translate("main_window", "Curricular Units 2nd Sem (without Evaluations)"))
        item = self.data_table.horizontalHeaderItem(31)
        item.setText(_translate("main_window", "Unemployment Rate"))
        item = self.data_table.horizontalHeaderItem(32)
        item.setText(_translate("main_window", "Inflation Rate"))
        item = self.data_table.horizontalHeaderItem(33)
        item.setText(_translate("main_window", "GDP"))
        item = self.data_table.horizontalHeaderItem(34)
        item.setText(_translate("main_window", "Target"))
        self.import_btn.setText(_translate("MainWindow", "Import"))
        self.btn_close.setText(_translate("MainWindow", "Close"))
        self.btn_export.setText(_translate("MainWindow", "Export "))
        self.label_link.setText(_translate("MainWindow", "TextLabel"))
        # self.label_tt.setText(_translate("MainWindow", "TextLabel"))
        self.btn_compare.setText(_translate("MainWindow", "Compare"))
        self.btn_predict_all.setText(_translate("MainWindow", "Predict All"))
        self.label_tt.setText(_translate("MainWindow", "TextLabel"))
        self.label.setText(_translate("MainWindow", "Shape of data:"))
        self.label_shape.setText(_translate("MainWindow", ""))

    def import_csv(self):
        try:
            root = tk.Tk()
            root.withdraw()
            file_path = filedialog.askopenfilename()
            self.label_link.setText(file_path)
            if file_path == '':
                pass
            else:
                extension = pathlib.Path(file_path).suffix
                if extension == '.csv' or extension == '.xlsx':
                    if file_path is not None:
                        if extension == '.csv':
                            self.data = pd.read_csv(file_path)
                        else:
                            self.data = pd.read_excel(file_path)
                        self.data_table.setRowCount(len(self.data))
                        for i in range(len(self.data)):
                            for j in range(35):
                                self.data_table.setItem(
                                    i, j, QtWidgets.QTableWidgetItem(str(self.data.iloc[i, j])))
                                if j == 34:
                                    self.list_data.append(self.data.iloc[i, j])
                                    self.detail_btn = QtWidgets.QPushButton(self)
                                    self.detail_btn.clicked.connect(self.predict_btn_click)
                                    self.detail_btn.setText("Predict")
                                    self.data_table.setCellWidget(i, j, self.detail_btn)
                                    self.label_shape.setText(machine_learning.shape(file_path, extension))

                else:
                    QMessageBox.information(self, 'System', 'Please import into excel file!', QMessageBox.Close)
        except:
            QMessageBox.information(self, 'System', 'Error import. Please try again!', QMessageBox.Close)
        else:
            self.btn_predict_all.setEnabled(True)

    def predict_btn_click(self):
        try:
            button = QtWidgets.qApp.focusWidget()
            index = self.data_table.indexAt(button.pos())
            table_model = self.data_table.model()

            d0_index = table_model.index(index.row(), 0)
            d0 = table_model.data(d0_index)

            d1_index = table_model.index(index.row(), 1)
            d1 = table_model.data(d1_index)

            d2_index = table_model.index(index.row(), 2)
            d2 = table_model.data(d2_index)

            d3_index = table_model.index(index.row(), 3)
            d3 = table_model.data(d3_index)

            d4_index = table_model.index(index.row(), 4)
            d4 = table_model.data(d4_index)

            d5_index = table_model.index(index.row(), 5)
            d5 = table_model.data(d5_index)

            d7_index = table_model.index(index.row(), 7)
            d7 = table_model.data(d7_index)

            d8_index = table_model.index(index.row(), 8)
            d8 = table_model.data(d8_index)

            d9_index = table_model.index(index.row(), 9)
            d9 = table_model.data(d9_index)

            d10_index = table_model.index(index.row(), 10)
            d10 = table_model.data(d10_index)

            d11_index = table_model.index(index.row(), 11)
            d11 = table_model.data(d11_index)

            d12_index = table_model.index(index.row(), 12)
            d12 = table_model.data(d12_index)

            d13_index = table_model.index(index.row(), 13)
            d13 = table_model.data(d13_index)

            d14_index = table_model.index(index.row(), 14)
            d14 = table_model.data(d14_index)

            d15_index = table_model.index(index.row(), 15)
            d15 = table_model.data(d15_index)

            d16_index = table_model.index(index.row(), 16)
            d16 = table_model.data(d16_index)

            d17_index = table_model.index(index.row(), 17)
            d17 = table_model.data(d17_index)

            d19_index = table_model.index(index.row(), 19)
            d19 = table_model.data(d19_index)

            d20_index = table_model.index(index.row(), 20)
            d20 = table_model.data(d20_index)

            d21_index = table_model.index(index.row(), 21)
            d21 = table_model.data(d21_index)

            d22_index = table_model.index(index.row(), 22)
            d22 = table_model.data(d22_index)

            d23_index = table_model.index(index.row(), 23)
            d23 = table_model.data(d23_index)

            d24_index = table_model.index(index.row(), 24)
            d24 = table_model.data(d24_index)

            d25_index = table_model.index(index.row(), 25)
            d25 = table_model.data(d25_index)

            d26_index = table_model.index(index.row(), 26)
            d26 = table_model.data(d26_index)

            d27_index = table_model.index(index.row(), 27)
            d27 = table_model.data(d27_index)

            d28_index = table_model.index(index.row(), 28)
            d28 = table_model.data(d28_index)

            d29_index = table_model.index(index.row(), 29)
            d29 = table_model.data(d29_index)

            d30_index = table_model.index(index.row(), 30)
            d30 = table_model.data(d30_index)

            d31_index = table_model.index(index.row(), 31)
            d31 = table_model.data(d31_index)

            d32_index = table_model.index(index.row(), 32)
            d32 = table_model.data(d32_index)

            d33_index = table_model.index(index.row(), 33)
            d33 = table_model.data(d33_index)

            list_data = [d0, d1, d2, d3, d4, d5, d7, d8, d9, d10, d11, d12, d13, d14, d15, d16, d17, d19, d20, d21, d22,
                         d23, d24, d25, d26, d27, d28, d29, d30, d31, d32, d33]
            for i in range(0, len(list_data)):
                if list_data[i] == '':
                    list_data[i] = 0
                if list_data[i] == 'Yes':
                    list_data[i] = 1
                elif list_data[i] == 'No':
                    list_data[i] = 0
                elif list_data[i] == 'Morning':
                    list_data[i] = 1
                elif list_data[i] == 'Evening':
                    list_data[i] = 0
                elif list_data[i] == 'Male':
                    list_data[i] = 1
                elif list_data[i] == 'Female':
                    list_data[i] = 0
            path = self.label_link.text()
            tt = self.label_tt.text()
            if tt == 'Random Forest Classifier':
                re = machine_learning.random_forest(path, list_data, int(self.max_depth), float(self.number))
            elif tt == 'Logistic Regression':
                list_data = list(map(float, list_data))
                re = machine_learning.logic_regression(path, list_data, float(self.number))
            elif tt == 'Linear Regression':
                re = machine_learning.liner(path, list_data, 0.1)
            elif tt == 'Decision Tree Classifier':
                re = machine_learning.decision_tree(path, list_data, float(self.number), int(self.max_depth))
            elif tt == 'K-Nearest Neighbors Classifier':
                list_data = list(map(float, list_data))
                re = machine_learning.knn(path, list_data, float(self.number), int(self.max_depth))

            if re[0] == 0:
                result = 'Drop out'
                self.data_table.removeCellWidget(index.row(), 34)
                self.data_table.setItem(
                    index.row(), 34, QtWidgets.QTableWidgetItem(result))
            else:
                result = 'Graduate'
                self.data_table.removeCellWidget(index.row(), 34)
                self.data_table.setItem(
                    index.row(), 34, QtWidgets.QTableWidgetItem(result))

        except:
            QMessageBox.information(self, 'System', 'Error Predict. Please try again!', QMessageBox.Close)

    def predict_all(self):
        try:
            row = self.data_table.rowCount()
            list_data = []
            for i in range(row):
                for j in range(34):
                    data = self.data_table.item(i, j).text()
                    list_data.append(data)
            list_result = list(cm.divide_chunks(list_data, 34))
            for i in list_result:
                i.pop(6)
                i.pop(18)
            path = self.label_link.text()
            tt = self.label_tt.text()
            if tt == 'Random Forest Classifier':
                # result = machine_learning.random_forest(path, list_data, 5, 0.1)
                list_predict = []
                for i in list_result:
                    re = machine_learning.random_forest(path, i, int(self.max_depth), float(self.number))
                    list_predict.append(re[0])
                self.fill_predict_value(list_predict, row)
            elif tt == 'Logistic Regression':
                # list_data = list(map(float, list_data))
                # result = machine_learning.logic_regression(path, list_data, 0.1)
                list_predict = []
                for i in list_result:
                    list_data = list(map(float, i))
                    re = machine_learning.logic_regression(path, list_data, float(self.number))
                    list_predict.append(re[0])
                self.fill_predict_value(list_predict, row)
            elif tt == 'Linear Regression':
                # result = machine_learning.liner(path, list_data, 0.1)
                list_predict = []
                for i in list_result:
                    re = machine_learning.liner(path, list_data, float(self.number))
                    list_predict.append(re[0])
                self.fill_predict_value(list_predict, row)
            elif tt == 'Decision Tree Classifier':
                # result = machine_learning.decision_tree(path, list_data, 0.1)
                list_predict = []
                for i in list_result:
                    list_data = list(map(float, i))
                    re = machine_learning.decision_tree(path, list_data, float(self.number), int(self.max_depth))
                    # re = machine_learning.decision_tree(path, list_data, float(self.number), int(self.max_depth))
                    list_predict.append(re[0])
                self.fill_predict_value(list_predict, row)
            elif tt == 'K-Nearest Neighbors Classifier':
                # list_data = list(map(float, list_data))
                # result = machine_learning.knn(path, list_data, 0.1, 5)
                list_predict = []
                for i in list_result:
                    list_data = list(map(float, i))
                    re = machine_learning.knn(path, list_data, float(self.number), int(self.max_depth))
                    list_predict.append(re[0])
                self.fill_predict_value(list_predict, row)
        except:
            QMessageBox.information(self, 'System', 'An error occurred while predicting all. Please try again!',
                                    QMessageBox.Close)
        else:
            self.btn_export.setEnabled(True)
            self.btn_compare.setEnabled(True)

    def fill_predict_value(self, list_predict, row):
        try:
            for i in range(len(list_predict)):
                if list_predict[i] == 0:
                    list_predict[i] = 'Dropout'
                else:
                    list_predict[i] = 'Graduate'
            self.list_data_predict = list_predict
            for i in range(row):
                for j in range(len(list_predict)):
                    if i == j:
                        self.data_table.removeCellWidget(i, 34)
                        self.data_table.setItem(
                            i, 34, QtWidgets.QTableWidgetItem(list_predict[j]))
        except:
            QMessageBox.information(self, 'System', 'An error occurred while filling value predict. Please try again!',
                                    QMessageBox.Close)

    def close(self):
        self.hide()

    def compare(self):
        try:
            re = round(accuracy_score(self.list_data, self.list_data_predict), 3) * 100
            QMessageBox.information(self, 'System', f'The accuracy compared to the actual result is: {re}%',
                                    QMessageBox.Close)
            self.accuracy = re
        except Exception as err:
            QMessageBox.information(self, 'System', 'An error occurred while comparing result. Please try again!\n'
                                                    f'{str(err)}',
                                    QMessageBox.Close)

    def export(self):
        try:
            columnHeaders = []
            for j in range(self.data_table.model().columnCount()):
                columnHeaders.append(self.data_table.horizontalHeaderItem(j).text())
            df = pd.DataFrame(columns=columnHeaders)
            for row in range(self.data_table.rowCount()):
                for col in range(self.data_table.columnCount()):
                    df.at[row, columnHeaders[col]] = self.data_table.item(row, col).text()
            wb = Workbook()
            wb = load_workbook('../Template/Template_predict_export.xlsx')
            ws1 = wb.worksheets[0]
            offset_row = 1
            offset_col = 0
            row = 1
            for row_data in dataframe_to_rows(df, index=False, header=False):
                col = 1
                for cell_data in row_data:
                    ws1.cell(row + offset_row, col + offset_col, cell_data)
                    col += 1
                row += 1
            if os.path.exists('../Output') == False:
                folder_path = '../Output'
                os.mkdir(folder_path)
            (d, m, y, h, mi, s) = cm.split_date_time()
            wb.save(
                f'../Output/predict_{d}_{m}_{y}_{h}_{mi}_{s}.xlsx')
        except:
            QMessageBox.information(self, 'System',
                                    'An error occurred while exporting to excel. Please try again later!',
                                    QMessageBox.Close)

    def data_from_table(self):
        try:
            row = self.data_table.rowCount()
            df = pd.DataFrame(
                columns=['Marital status', 'Application mode', 'Application order', 'Course',
                         'Daytime/evening attendance', 'Previous qualification', 'Nacionality',
                         'Mother\'s qualification', 'Father\'s qualification',
                         'Mother\'s occupation', 'Father\'s occupation', 'Displaced',
                         'Educational special needs', 'Debtor', 'Tuition fees up to date',
                         'Gender', 'Scholarship holder', 'Age at enrollment', 'International',
                         'Curricular units 1st sem (credited)',
                         'Curricular units 1st sem (enrolled)',
                         'Curricular units 1st sem (evaluations)',
                         'Curricular units 1st sem (approved)',
                         'Curricular units 1st sem (grade)',
                         'Curricular units 1st sem (without evaluations)',
                         'Curricular units 2nd sem (credited)',
                         'Curricular units 2nd sem (enrolled)',
                         'Curricular units 2nd sem (evaluations)',
                         'Curricular units 2nd sem (approved)',
                         'Curricular units 2nd sem (grade)',
                         'Curricular units 2nd sem (without evaluations)', 'Unemployment rate',
                         'Inflation rate', 'GDP', 'Target'])
            for i in range(row):
                value = []
                for j in range(35):
                    value.append(self.data_table.item(i, j).text())
                df.loc[len(df.index)] = value
            return df
        except Exception as err:
            QMessageBox.information(self, 'Error', str(err.args[0]), QMessageBox.Close)
            return []

    def export_button_active(self):
        try:
            tt = self.label_tt.text()
            link_predict = self.link_export = cm.get_path('predict')
            path = self.label_link.text()
            df = self.data_from_table()
            if len(df.index) == 0:
                QMessageBox.information(self, 'Error', 'There is no value', QMessageBox.Close)
                return
            if cm.data_to_excel('predict', df):
                machine_learning.percentage_of_student_target(path, 'export')
                # machine_learning.number_of_student_target_by_genders(path,'export')
                machine_learning.number_of_student_target_by_arital_status(path, 'export')
                machine_learning.number_of_student_enrolled_by_ages(path, 'export')
                machine_learning.dropout_percentage_by_ges(path, 'export')
                machine_learning.number_of_student_with_educational_special_needs(path, 'export')
                machine_learning.statistics_comparison_in_number_of_studen_target_by_debt(path, 'export')
                machine_learning.rate_of_unemployment_student(path, 'export')
                machine_learning.correlation_heatmap_between_variables(path, 'export')
                workbook = xlsxwriter.Workbook(link_predict)
                data = pd.read_excel(link_predict)
                worksheet = workbook.add_worksheet('Predict data')
                worksheet_2 = workbook.add_worksheet('Graph')
                worksheet_3 = workbook.add_worksheet('Report')
                bold = workbook.add_format({'bold': True})
                red = workbook.add_format({'font_color': 'red'})
                bold_red = workbook.add_format({'font_color': 'red', 'bold': True, 'font_size': 13})
                merge_format = workbook.add_format({'align': 'center', 'align': 'center', 'valign': 'vcenter'})
                bold_centered_format = workbook.add_format({'bold': True, 'align': 'center', 'font_size': 14})

                for i, col_name in enumerate(data.columns):
                    worksheet.write(0, i, col_name, bold)
                    worksheet.write_column(1, i, data[col_name])
                worksheet_2.set_column("A:A", 30)
                worksheet_2.write("A1", "Percentage of Student Target:", bold_red)
                worksheet_2.insert_image("A2", "graph1.png", {"x_scale": 0.7, "y_scale": 0.7})

                worksheet_2.write("F1", "Number of student target by Genders:", bold_red)
                worksheet_2.insert_image("F2", "graph3.png", {"x_scale": 0.7, "y_scale": 0.7})

                worksheet_2.write("P1", "Number of student target by Marital Status:", bold_red)
                worksheet_2.insert_image("P2", "graph4.png", {"x_scale": 0.7, "y_scale": 0.7})

                worksheet_2.write("A21", "Dropout percentage by Ages:", bold_red)
                worksheet_2.insert_image("A22", "graph5.png", {"x_scale": 0.7, "y_scale": 0.7})

                worksheet_2.write("F21", "Number of student with Educational Special Needs:", bold_red)
                worksheet_2.insert_image("F22", "graph6.png", {"x_scale": 0.7, "y_scale": 0.7})

                worksheet_2.write("P21", "Statistics/comparison in Number of student Target by Debt:", bold_red)
                worksheet_2.insert_image("P22", "graph7.png", {"x_scale": 0.7, "y_scale": 0.7})

                worksheet_2.write("A42", "Rate of Unemployment student:", bold_red)
                worksheet_2.insert_image("A43", "graph8.png", {"x_scale": 0.7, "y_scale": 0.7})

                # worksheet_2.write("F42", "Correlation heatmap between variables:")
                # worksheet_2.insert_image("F43", "graph9.png", {"x_scale": 0.7, "y_scale": 0.7})

                worksheet_2.write("A62", "Assess the accuracy of the model through the confusion matrix scale:",
                                  bold_red)
                worksheet_2.insert_image("A63", "graph9.png", {"x_scale": 0.7, "y_scale": 0.7})
                # f1 = worksheet_3.add_format({'bold': True, 'font_color': 'black'})
                worksheet_3.merge_range(0, 0, 0, 9, ' ', merge_format)
                worksheet_3.write('A1', 'REPORT', bold_centered_format)
                worksheet_3.write('A2', 'No.', bold)
                worksheet_3.write('B2', 'Predict', bold)
                worksheet_3.write('C2', 'Reality', bold)
                list_num = cm.createList(len(self.list_data))
                list_num_to_str = list(map(str, list_num))

                for i, col_name in enumerate(list_num_to_str):
                    worksheet_3.write(2, 0, col_name)
                    worksheet_3.write_column(2, 0, list_num_to_str)
                for i, col_name in enumerate(self.list_data_predict):
                    worksheet_3.write(2, 1, col_name)
                    worksheet_3.write_column(2, 1, self.list_data_predict)
                for i, col_name in enumerate(self.list_data):
                    worksheet_3.write(2, 2, col_name)
                    worksheet_3.write_column(2, 2, self.list_data)
                number = len(self.list_data)
                correct = cm.compare(self.list_data, self.list_data_predict)
                worksheet_3.merge_range(number + 4, 0, number + 4, 9, '', merge_format)
                worksheet_3.write(number + 4, 0,
                                  f'The algorithm correctly predicted: {correct}/{number} compared to the reality.')
                if tt == 'Random Forest Classifier' or tt == 'Decision Tree Classifier':
                    worksheet_3.merge_range(number + 3, 0, number + 3, 9, '', merge_format)
                    worksheet_3.write(number + 3, 0,
                                      f'The algorithm used is: {tt} (test_size={self.number}, max_depth={self.max_depth})')
                elif tt == 'Logistic Regression':
                    worksheet_3.merge_range(number + 3, 0, number + 3, 9, '', merge_format)
                    worksheet_3.write(number + 3, 0,
                                      f'The algorithm used is: {tt} (test_size={self.number}, max_depth={self.max_depth})')
                else:
                    worksheet_3.merge_range(number + 3, 0, number + 3, 9, '', merge_format)
                    worksheet_3.write(number + 3, 0,
                                      f'The algorithm used is: {tt} (test_size={self.number}, n_neighbors={self.max_depth})')
                worksheet_3.merge_range(number + 5, 0, number + 5, 9, '', merge_format)
                worksheet_3.write(number + 5, 0, f'The algorithm has an accuracy of: {self.accuracy}%', red)

                workbook.close()
                os.remove('graph1.png')
                # os.remove('graph2.png')
                os.remove('graph3.png')
                os.remove('graph4.png')
                os.remove('graph5.png')
                os.remove('graph6.png')
                os.remove('graph7.png')
                os.remove('graph8.png')
                os.remove('graph9.png')
                QMessageBox.information(self, 'Success', 'Export success full', QMessageBox.Close)
            else:
                QMessageBox.information(self, 'Fail', 'Something wrong', QMessageBox.Close)
        except Exception as err:
            QMessageBox.information(self, 'Error', str(err), QMessageBox.Close)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ui = UI_advanced_prediction()
    ui.setupUi()
    ui.show()
    sys.exit(app.exec_())
